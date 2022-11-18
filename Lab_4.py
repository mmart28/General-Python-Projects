#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#Madeline Martine

import openpyxl

manifest=open('W:\MGMT58600\manifest.txt','a') #open manifest to add data

print('Do you need to enter any cars not already included in the manifest? (Y/N)') #ask user if they have additional car data to add
answer=input()

while answer=='Y': #while loop will run and collect car data as long as user has more car data to add to manifest
    print('Please enter the name of the car:')
    name=input()
    print('Please enter the MSRP:')
    MSRP=input()
    print('Please enter KSMs cost:')
    KSMcost=input()
    print('Please enter the code for the car (0=domestic, 1=import):')
    code=input()
    print('Please enter the code for the commission rate (A/B/C):')
    rate=input()
    newline='\n'+name+' '+MSRP+' '+KSMcost+' '+code+' '+rate #formatting car data to be entered into manifest
    manifest.write(newline) #write new car data to the manifest
    print('Do you have another car to enter? (Y/N)') #user response will determine if while loop runs again
    answer=input()
    
manifest.close() #close the manifest

manifest=open('W:\MGMT58600\manifest.txt')
manifest_data=manifest.read() #opening the manifest file and storing the data

list_data=[]
list_data=manifest_data.split('\n') #storing the car data in a list and seperating the info by line

def commission(MSRP,KSMcost,code,rate):
    '''Function calculates the potential gross profit and adjusts based on entered code (domestic or import) 
    then returns expected commission based on entered rate (A/B/C) that determines percentage'''
    if code=='0':
        AdjPGP=KSMcost-MSRP
        if rate=='A':
            return AdjPGP*.35
        elif rate=='B':
            return AdjPGP*.25
        elif rate=='C':
            return AdjPGP*.15
    elif code=='1':
        AdjPGP=(KSMcost-MSRP)*.9825
        if rate=='A':
            return AdjPGP*.35
        elif rate=='B':
            return AdjPGP*.25
        elif rate=='C':
            return AdjPGP*.15

wb=openpyxl.load_workbook('W:\MGMT58600\commission.xlsx') #opening excel sheet
sheet=wb['commission'] #telling the program which excel sheet to write to

#the for loop will go through each item in the list (each line containing info for 1 car) and break it up into its parts:
#name, MSRP, KSM cost, code, and rate. Then those variables are put into the commission function to determine expected
#commission for each car
for i in range(len(list_data)):
    breakup=list_data[i].split()
    name=breakup[0]
    MSRP=float(breakup[1])
    KSMcost=float(breakup[2])
    code=breakup[3]
    rate=breakup[4]
    pot_commission=commission(MSRP,KSMcost,code,rate)
    sheet.cell(row=i+1,column=1).value=name #car name is being put into col1 of the excel sheet
    sheet.cell(row=i+1,column=2).value=round(pot_commission,2) #calculated commission is being put into col2 of the excel sheet
wb.save('W:\MGMT58600\commission.xlsx') #making sure the data in the excel sheet is saved

#notify the user that the program has completed and data has been stored in manifest and excel sheet
print('\nThe manifest has been updated and potential commissions have been entered into the Excel sheet.')


# In[ ]:




